import openpyxl as px
from shutil import copyfile
from os import path
import sys
from analyzeGPA_okadai import AnalyzeGPA_Okadai
from main import open_gradefile


class ExportToExcelFile:
    def __init__(self, dirname=None):
        if dirname is None:
            dirname = path.join(path.dirname(path.abspath(sys.argv[0])),
                                "../excel_file/")
        updated_date, gpa_list = open_gradefile()
        if updated_date is not False:
            self.analyzegpa_okadai = AnalyzeGPA_Okadai(gpa_list)
            self.excel_filename =\
                ("gpaAnalyzer_{}"
                 ".xlsx".format(updated_date.strftime('%Y%m%d')))
            self.excel_filename = path.join(dirname, self.excel_filename)
            self.original_excel_filename =\
                path.join(path.dirname(path.abspath(sys.argv[0])),
                          "../excel_file/gpaAnalyzer.xlsx")

    def export_to_excel_file(self):
        try:
            wb = px.load_workbook(self.original_excel_filename)
            ws_gpa_okadai = wb['Okadai']
            ws_gpa_summary = wb['Summary']

            year_completed_list, gpa_dict = self.analyzegpa_okadai.get_gpa()
            for i, yc in enumerate(year_completed_list):
                ws_gpa_okadai.cell(row=i+2, column=1, value=yc)
                ws_gpa_okadai.cell(row=i+2, column=2, value=gpa_dict[yc]["gpa"])
                ws_gpa_okadai.cell(row=i+2, column=3,
                                   value=gpa_dict[yc]["accumulated_gpa"])
            ws_gpa_okadai.row_dimensions.group(i+3, 32, hidden=True)
            wb.save(self.excel_filename)
        finally:
            wb.close()


if __name__ == "__main__":
    etef = ExportToExcelFile()
    etef.export_to_excel_file()
